from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from modules.audit_engine.core.field_resolver import resolve_all_fields
from modules.audit_engine.services.mapping_service import map_project_name
from modules.audit_engine.services.business_excel_package_parser import try_parse_business_package
from modules.audit_engine.services.excel_row_mapper import map_excel_row_to_field_candidates


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def _is_present(value: Any) -> bool:
    return value is not None and value != ""


def _build_headers(values: Tuple[Any, ...]) -> Tuple[List[str], List[str]]:
    headers: List[str] = []
    warnings: List[str] = []
    seen: Dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        header = str(value or "").strip()
        if not header:
            header = f"未命名列{index}"
            warnings.append(f"第 {index} 列表头为空，已按 {header} 处理。")

        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            unique_header = f"{header}__{count + 1}"
            warnings.append(f"表头 {header} 重复，重复列已标记为 {unique_header}。")
            header = unique_header
        headers.append(header)

    return headers, warnings


def _parse_flat_table_workbook(workbook: Any, filename: str = "") -> Dict[str, Any]:
    sheet = workbook.worksheets[0]
    warnings: List[str] = []
    rows: List[Dict[str, Any]] = []

    row_iter = sheet.iter_rows(values_only=True)
    try:
        header_values = next(row_iter)
    except StopIteration:
        workbook.close()
        return {
            "filename": filename,
            "file_type": "xlsx",
            "status": "parsed",
            "parse_mode": "flat_table",
            "sheet_name": sheet.title,
            "rows": [],
            "documents": [],
            "warnings": ["Excel 文件为空。"],
        }

    headers, header_warnings = _build_headers(header_values)
    warnings.extend(header_warnings)

    for excel_row_index, values in enumerate(row_iter, start=2):
        row = {
            header: _cell_value(values[index]) if index < len(values) else None
            for index, header in enumerate(headers)
        }
        if not any(_is_present(value) for value in row.values()):
            continue

        mapped = map_excel_row_to_field_candidates(row, filename=filename, sheet_name=sheet.title)
        resolved = resolve_all_fields(mapped.get("field_candidates") or {}, catalog_mapper=map_project_name)
        project_name = str(resolved["standard_fields"].get("project_name", {}).get("value") or "")
        audit_request = {
            "project_name": project_name,
            "standard_fields": resolved["standard_fields"],
            "missing_fields": resolved["missing_fields"],
            "conflicting_fields": resolved["conflicting_fields"],
            "warnings": resolved["warnings"],
            "mapped_objects": resolved["mapped_objects"],
            "matched_object_ids": resolved["matched_object_ids"],
        }
        rows.append(
            {
                "row_index": excel_row_index,
                "project_key": str(excel_row_index),
                "project_name": project_name,
                "raw_row": row,
                "standard_fields": resolved["standard_fields"],
                "missing_fields": resolved["missing_fields"],
                "conflicting_fields": resolved["conflicting_fields"],
                "audit_ready": resolved["audit_ready"],
                "audit_request": audit_request,
                "source_sheets": [sheet.title],
                "business_summary": ["已按扁平表模式解析；每行作为一个项目审计。"],
                "warnings": resolved["warnings"],
                "mapped_objects": resolved["mapped_objects"],
                "matched_object_ids": resolved["matched_object_ids"],
                "debug": {
                    "source_sheets": [sheet.title],
                    "unmapped_columns": mapped.get("unmapped_columns") or [],
                },
            }
        )

    return {
        "filename": filename,
        "file_type": "xlsx",
        "status": "parsed",
        "parse_mode": "flat_table",
        "sheet_name": sheet.title,
        "rows": rows,
        "documents": [],
        "warnings": warnings,
        "business_summary": ["已按扁平表模式解析；每行作为一个项目审计。"],
    }


def parse_xlsx_bytes(content: bytes, filename: str = "") -> Dict[str, Any]:
    """Parse one .xlsx file into project-level audit requests.

    Business export workbooks are detected and aggregated across sheets first.
    If a workbook is not a business package, the parser falls back to the flat
    table mode where each row is treated as one audit project.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法解析 .xlsx 文件。") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        business_result = try_parse_business_package(workbook, filename=filename)
        if business_result:
            return business_result
        return _parse_flat_table_workbook(workbook, filename=filename)
    finally:
        workbook.close()
